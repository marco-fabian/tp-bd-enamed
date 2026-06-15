#!/usr/bin/env python3
"""
ETL — constrói data/processed/enamed.db a partir dos microdados reais do
Enamed 2025 (INEP).

Decisões de modelagem (ver Seção 3 do relatório):
- A população usada é a do ENADE (TP_INSCRICAO = 1): são os únicos
  participantes que possuem, ao mesmo tempo, curso/IES/localização (arq1)
  e desempenho/caderno/vetores (arq3) — exatamente o que o modelo unificado
  de Estudante exige.
- Os arquivos do INEP são particionados por COLUNA e NÃO estão alinhados por
  linha: o único elo entre eles é CO_CURSO (nível de curso). Por isso:
    * curso/IES/município/UF        <- arq1  (dimensão, por CO_CURSO)
    * caderno/vetores/notas/presença<- arq3  (fato, por estudante)
    * itens (TRI)                   <- parametros_itens
  Atributos demográficos (idade, sexo, ano de conclusão, turno) estão em
  arq2/arq5/arq6, em ordem diferente, sem chave de estudante — logo NÃO podem
  ser ligados ao estudante correto e ficam NULL em Estudante.
- IDH (Municipio), MODALIDADE e QTD_VAGAS (Curso) não existem nos microdados
  (precisariam de fontes externas: Atlas Brasil, e-MEC) e ficam NULL.
"""
import os
import csv
import sqlite3
import glob
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DDL = os.path.join(ROOT, "sql", "01_create_tables.sql")
DB = os.path.join(ROOT, "data", "processed", "enamed.db")
EXTRACTED = os.path.join(ROOT, "data", "extracted")
RAW = os.path.join(ROOT, "data", "raw")

# ----------------------------------------------------------------- mapas IBGE
UF_NOME = {
    11: "Rondônia", 12: "Acre", 13: "Amazonas", 14: "Roraima", 15: "Pará",
    16: "Amapá", 17: "Tocantins", 21: "Maranhão", 22: "Piauí", 23: "Ceará",
    24: "Rio Grande do Norte", 25: "Paraíba", 26: "Pernambuco", 27: "Alagoas",
    28: "Sergipe", 29: "Bahia", 31: "Minas Gerais", 32: "Espírito Santo",
    33: "Rio de Janeiro", 35: "São Paulo", 41: "Paraná", 42: "Santa Catarina",
    43: "Rio Grande do Sul", 50: "Mato Grosso do Sul", 51: "Mato Grosso",
    52: "Goiás", 53: "Distrito Federal",
}
REGIAO_NOME = {1: "Norte", 2: "Nordeste", 3: "Sudeste", 4: "Sul", 5: "Centro-Oeste"}


def achar_dados():
    """Localiza a pasta DADOS; extrai o zip se necessário."""
    hits = glob.glob(os.path.join(EXTRACTED, "**", "DADOS"), recursive=True)
    if not hits:
        zips = glob.glob(os.path.join(RAW, "*.zip"))
        if not zips:
            sys.exit("Coloque o zip dos microdados em data/raw/ (ou extraia em data/extracted/).")
        os.makedirs(EXTRACTED, exist_ok=True)
        print("Extraindo", os.path.basename(zips[0]), "...")
        import zipfile
        with zipfile.ZipFile(zips[0]) as z:
            z.extractall(EXTRACTED)
        hits = glob.glob(os.path.join(EXTRACTED, "**", "DADOS"), recursive=True)
    return hits[0]


def num(s):
    """'' -> None ; '-3.655' -> -3.655 ; troca ',' por '.'."""
    s = (s or "").strip().replace(",", ".")
    if s == "" or s == ".":
        return None
    try:
        return float(s) if ("." in s) else int(s)
    except ValueError:
        return None


def carregar_municipios(dados_dir):
    """Lê a aba MUNICÍPIOS do dicionário -> {codigo: nome}."""
    leiame = os.path.join(os.path.dirname(dados_dir), "1. LEIA-ME")
    xlsx = [x for x in glob.glob(os.path.join(leiame, "Dicion*.xlsx")) if "~$" not in x]
    nomes = {}
    if not xlsx:
        print("  (dicionário de municípios não encontrado — nomes ficarão N/D)")
        return nomes
    from openpyxl import load_workbook
    wb = load_workbook(xlsx[0], read_only=True, data_only=True)
    if "MUNICÍPIOS" in wb.sheetnames:
        ws = wb["MUNICÍPIOS"]
        for row in ws.iter_rows(values_only=True):
            # a planilha tem 1ª coluna vazia: código em row[1], nome em row[2]
            if not row or len(row) < 3 or row[1] is None:
                continue
            try:
                cod = int(row[1])
            except (ValueError, TypeError):
                continue
            nome = str(row[2]).strip().title() if row[2] else None
            if nome:
                nomes[cod] = nome
    print(f"  municípios no dicionário: {len(nomes)}")
    return nomes


def main(com_resposta=False):
    dados = achar_dados()
    enade = os.path.join(dados, "Enade")
    arq1 = os.path.join(enade, "microdados_enade_2025_arq1.txt")
    arq3 = os.path.join(enade, "microdados_enade_2025_arq3.txt")
    params = os.path.join(dados, "microdados2025_parametros_itens.txt")

    os.makedirs(os.path.dirname(DB), exist_ok=True)
    if os.path.exists(DB):
        os.remove(DB)
    con = sqlite3.connect(DB)
    with open(DDL, encoding="utf-8") as f:
        con.executescript(f.read())
    con.execute("PRAGMA foreign_keys = OFF")  # inserimos dimensões antes dos fatos

    muni_nomes = carregar_municipios(dados)

    # ---------- arq1 -> dimensões Curso / IES / Municipio / UF ----------
    cursos, ies, munis, ufs = {}, {}, {}, {}
    with open(arq1, encoding="utf-8-sig") as f:
        r = csv.DictReader(f, delimiter=";")
        for d in r:
            co_curso = num(d["CO_CURSO"])
            co_ies = num(d["CO_IES"])
            co_muni = num(d["CO_MUNIC_CURSO"])
            co_uf = num(d["CO_UF_CURSO"])
            co_reg = num(d["CO_REGIAO_CURSO"])
            if co_curso is None:
                continue
            cursos.setdefault(co_curso, (co_ies, co_muni))
            if co_ies is not None:
                ies.setdefault(co_ies, (num(d["CO_CATEGAD"]), num(d["CO_ORGACAD"])))
            if co_muni is not None:
                munis.setdefault(co_muni, co_uf)
            if co_uf is not None:
                ufs.setdefault(co_uf, co_reg)

    con.executemany("INSERT INTO UF VALUES (?,?,?)",
                    [(uf, UF_NOME.get(uf, f"UF {uf}"), REGIAO_NOME.get(reg, "Centro-Oeste"))
                     for uf, reg in ufs.items()])
    con.executemany("INSERT INTO Municipio VALUES (?,?,?,?)",
                    [(m, muni_nomes.get(m, "N/D"), None, uf) for m, uf in munis.items()])
    con.executemany("INSERT INTO IES VALUES (?,?,?)",
                    [(i, cat, org) for i, (cat, org) in ies.items()])
    con.executemany("INSERT INTO Curso VALUES (?,?,?,?,?)",
                    [(c, None, None, i, m) for c, (i, m) in cursos.items()])
    print(f"  UF={len(ufs)}  Municipio={len(munis)}  IES={len(ies)}  Curso={len(cursos)}")

    # ---------- parametros_itens -> Item_prova + Composicao + Caderno ----------
    itens, comp = [], []
    with open(params, encoding="utf-8-sig") as f:
        r = csv.DictReader(f, delimiter=";")
        for k, d in enumerate(r, start=1):
            itens.append((k, num(d["OUTFIT"]), num(d["INFIT"]),
                          num(d["COR_BISSERIAL"]), num(d["PARAMETRO_B"]),
                          num(d["ITEM_MANTIDO"])))
            p1, p2 = num(d["NU_ITEM_PROVA_1"]), num(d["NU_ITEM_PROVA_2"])
            if p1 is not None:
                comp.append((1, k, p1))
            if p2 is not None:
                comp.append((2, k, p2))
    con.executemany("INSERT INTO Item_prova VALUES (?,?,?,?,?,?)", itens)

    # Caderno (NU_ITEM/X/Z são constantes por caderno; pega do arq3 depois)
    cadernos = {}

    # ---------- arq3 -> Estudante / Notas / Vetores ----------
    estudantes, notas, vetores = [], [], []
    idx = 0
    cursos_validos = set(cursos)
    with open(arq3, encoding="utf-8-sig") as f:
        r = csv.DictReader(f, delimiter=";")
        for d in r:
            idx += 1
            cad = num(d["CO_CADERNO"]) or 1
            cadernos.setdefault(cad, (num(d["NU_ITEM"]), num(d["NU_ITEM_X"]), num(d["NU_ITEM_Z"])))
            co_curso = num(d["CO_CURSO"])
            if co_curso not in cursos_validos:
                # curso ausente na dimensão: cria stub p/ não quebrar a FK
                con.execute("INSERT OR IGNORE INTO Curso VALUES (?,?,?,?,?)",
                            (co_curso, None, None, None, None))
                cursos_validos.add(co_curso)
            notas.append((idx,
                          num(d["QT_ACERTO_AREA_1"]), num(d["QT_ACERTO_AREA_2"]),
                          num(d["QT_ACERTO_AREA_3"]), num(d["QT_ACERTO_AREA_4"]),
                          num(d["QT_ACERTO_AREA_5"]),
                          num(d["PROFICIENCIA"]), num(d["NT_GER"]),
                          num(d["PER_ACERTO_ENARE"])))
            vetores.append((idx, num(d["NU_ANO"]),
                            d["DS_VT_GAB_OBJ"], d["DS_VT_ACE_OBJ"], d["DS_VT_ESC_OBJ"]))
            estudantes.append((idx, "1", None, None, None, None, None,
                               d["TP_PRES"].strip() or None, co_curso, cad, idx, idx))

    con.executemany("INSERT INTO Caderno VALUES (?,?,?,?)",
                    [(c, ni, nx, nz) for c, (ni, nx, nz) in cadernos.items()])
    # Composicao só p/ cadernos existentes
    comp = [c for c in comp if c[0] in cadernos]
    con.executemany("INSERT INTO Composicao VALUES (?,?,?)", comp)
    con.executemany("INSERT INTO Notas VALUES (?,?,?,?,?,?,?,?,?)", notas)
    con.executemany("INSERT INTO Vetores VALUES (?,?,?,?,?)", vetores)
    con.executemany(
        "INSERT INTO Estudante VALUES (?,?,?,?,?,?,?,?,?,?,?,?)", estudantes)
    print(f"  Caderno={len(cadernos)}  Item_prova={len(itens)}  Composicao={len(comp)}")
    print(f"  Estudante={len(estudantes)}  Notas={len(notas)}  Vetores={len(vetores)}")

    # ---------- Resposta (opcional, pesado: ~estudantes x itens) ----------
    if com_resposta:
        print("  gerando Resposta (pode demorar)...")
        pos2item = {}  # (caderno, posicao) -> item
        for cad, item, pos in comp:
            pos2item[(cad, pos)] = item
        ins = []
        for est in estudantes:
            co_vet = est[11]
            cad = est[9]
        # nota: reconstrução completa fica como extensão; tabela criada vazia
    con.commit()
    con.close()
    print("OK ->", DB)


if __name__ == "__main__":
    main(com_resposta="--resposta" in sys.argv)
